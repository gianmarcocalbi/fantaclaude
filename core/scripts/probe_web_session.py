"""One-shot discovery of the fantacalcio.it website session and the voti workbook.

Run from the workspace root, once, with FANTACALCIO_WEB_COOKIE in .env:

    uv run python core/scripts/probe_web_session.py

Four GETs against www.fantacalcio.it -- the current season's giornata 1, two
back seasons' giornata 1, and a giornata not played yet -- one second apart,
then a read-only look at whatever came back: status and size per URL, sheet
names, the first rows of every sheet, and the share of player codes the
current listone knows. It never prints the cookie. Not a test and not an
adapter: ingest/stats_web.py is written against what this prints, and the
spec's open question 5 records it.
"""

from __future__ import annotations

import asyncio
import sys

import duckdb
import openpyxl
from fantaclaude.config import web_cookie
from fantaclaude.ingest.http import (
    NotPublished,
    SourceError,
    WebSessionExpired,
    build_http,
    fetch_bytes,
    polite_pause,
)
from fantaclaude.paths import db_path, workspace_root

VOTES = "https://www.fantacalcio.it/api/v1/Excel/votes/{season}/{giornata}"
PROBES = [(21, 1), (20, 1), (18, 1), (21, 38)]      # season ids: 21 is 2026-27 (model/seasons.py)
XLSX_MAGIC = b"PK\x03\x04"


async def main() -> int:
    cookie = web_cookie()
    if cookie is None:
        print("FANTACALCIO_WEB_COOKIE is not set in .env -- capture it first (plan, Task 6, Step 1)")
        return 3
    captured = workspace_root() / "captured"
    captured.mkdir(exist_ok=True)
    saved = []
    async with build_http() as http:
        for index, (season, giornata) in enumerate(PROBES):
            if index:
                await polite_pause()
            url = VOTES.format(season=season, giornata=giornata)
            try:
                data = await fetch_bytes(http, url, headers={"Cookie": cookie})
            except WebSessionExpired as exc:
                print(f"{url}: session rejected (HTTP {exc.status}) -- the cookie does not authenticate; re-capture it")
                return 3
            except NotPublished:
                print(f"{url}: HTTP 404 -- not published")
                continue
            except SourceError as exc:
                print(f"{url}: {exc}")
                continue
            head = data[:2000].lower()
            kind = "xlsx" if data[:4] == XLSX_MAGIC else ("html" if b"<html" in head else "unknown")
            print(f"{url}: HTTP 200, {len(data)} bytes, {kind}")
            if kind == "xlsx":
                path = captured / f"voti-{season}-{giornata:02d}.xlsx"
                path.write_bytes(data)
                saved.append(path)
    known: set[int] = set()
    if db_path().is_file():
        con = duckdb.connect(str(db_path()), read_only=True)
        known = {int(r[0]) for r in con.execute("SELECT player_id FROM v_players_current").fetchall()}
        con.close()
    for path in saved:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print(f"\n== {path.name}: sheets {workbook.sheetnames}")
        for sheet in workbook.worksheets:
            print(f"-- {sheet.title}")
            codes: list[int] = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index < 8:
                    print("  ", [cell for cell in row if cell is not None])
                if row and isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
                    codes.append(int(row[0]))
            if codes:
                share = (sum(code in known for code in codes) / len(codes)) if known else 0.0
                print(f"   {len(codes)} player rows; {share:.0%} of the codes are in the current listone")
        workbook.close()
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
