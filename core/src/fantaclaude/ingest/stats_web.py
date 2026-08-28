"""Per-giornata voti and event counts from fantacalcio.it's XLSX export.

GET /api/v1/Excel/votes/<season_id>/<giornata>, sent the *website* cookie
from .env (captured by the account holder, never obtained by code). One
workbook per giornata; every sheet is one voto source (Redazione
Fantacalcio, Statistico, Italia) and is kept under its sheet name; rows are
grouped by club, each block opened by a row that carries only the club
name; a senza-voto is "6*" / "s.v." / the sentinel 55. `Cod.` is the
fantacalcio.it player id -- the listone's `id` -- so nothing here is
matched by name. The layout constants below are what Task 6 of the
Phase 0b plan observed; a header that differs is a red ingest, never a
silently-null column. Base voto and event counts only: the fantavoto is
computed at projection time under the league's own bonus/malus.

A giornata that has not been played yet answers HTTP 200 with a placeholder
workbook, not a 404 -- one sheet named "Fantacalcio" containing exactly one
row whose single cell reads "File ancora non disponibile. Riprova più
tardi" (observed 2026-08-28, captured/voti-21-38.xlsx). fetch_voti treats
that the same as a 404: NotPublished, and nothing is written to data/raw/.

Each club block ends with one extra row for the coach, whose `Ruolo` is
"ALL" (e.g. `688 'ALL' 'Sarri'`, `captured/voti-21-01.xlsx`). It has a
voto like any player row and is stored the same way -- dropping it would
hardcode "this league does not score the coach", which not every league
agrees with. A coach id is never in the listone, so it is excluded from
`unknown_players` (that count measures listone coverage of *players*); a
consumer that wants players only filters `classic_role <> 'ALL'` --
Phase 1's projection does this (plan, Task 7).
"""

from __future__ import annotations

import io
import json
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import duckdb
import httpx
import openpyxl

from fantaclaude.ingest.http import (
    NotPublished,
    WebSessionExpired,
    fetch_bytes,
    polite_pause,
)
from fantaclaude.ingest.raw import RawFile, RawStore
from fantaclaude.timeutil import to_db

SOURCE = "fantacalcio.it:/api/v1/Excel/votes"
VOTES_URL = "https://www.fantacalcio.it/api/v1/Excel/votes/{season_id}/{giornata}"
XLSX_MAGIC = b"PK\x03\x04"
ACCEPT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*"
# The header row of every sheet, as observed in Task 6 -- fail loud on anything else.
VOTI_HEADER = ("Cod.", "Ruolo", "Nome", "Voto", "Gf", "Gs", "Rp", "Rs", "Rf", "Au", "Amm", "Esp", "Ass")
COLUMNS = {"Gf": "goals", "Gs": "goals_conceded", "Rp": "pen_saved", "Rs": "pen_missed", "Rf": "pen_scored",
           "Au": "own_goals", "Amm": "yellow", "Esp": "red", "Ass": "assists"}
SENZA_VOTO_TEXT = frozenset({"s.v.", "s.v", "sv", "-", ""})
SENZA_VOTO_SENTINEL = Decimal(55)          # the voti page's data-value for a player without a voto
HEADER_SEARCH_ROWS = 20
# Ruling R6 (Task 7 dispatch): the coach row that closes every club block --
# stored like any other row, but never in the listone, so excluded from
# unknown_players (which measures listone coverage of players).
COACH_ROLE = "ALL"
# Ruling R4 (Task 7 dispatch): the site's own text for a giornata not yet
# published, observed verbatim in captured/voti-21-38.xlsx -- a substring
# match, since it is the whole (and only) cell of the placeholder's one row.
NOT_PUBLISHED_MARKER = "File ancora non disponibile"


class VotiShapeError(ValueError):
    """The workbook is not the voti export this adapter was written against."""


def _is_not_published_placeholder(data: bytes) -> bool:
    """True when `data` is the "not yet published" placeholder workbook, not
    a genuine voti export -- checked from the raw bytes before anything is
    stored, so the adapter never persists a placeholder as if it were data.

    Called only once `data` has already passed the xlsx magic-byte check, so
    a load failure here is a genuine shape surprise and is left to propagate
    -- same as parse_voti, which does not guard load_workbook either."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and NOT_PUBLISHED_MARKER in cell:
                        return True
        return False
    finally:
        workbook.close()


# Ruling R9 (Task 8): a giornata on the calendar that has not been rated yet
# is a third shape, distinct from both the placeholder above and a genuine
# voti export -- observed 2026-08-29, captured/voti-21-03.xlsx (season 21's
# giornata 3, two rounds into a season whose auction has not happened yet).
# Every sheet carries only the title-and-disclaimer block every workbook
# opens with, and then nothing else: no club row, no header, no player row.
# Matched narrowly, exactly against that fixed block: any extra row, in any
# sheet -- a header, a club, a player, or anything unrecognised -- and this
# is not it, so parse_voti's "no sheet carries the voti table" keeps failing
# loud on a genuine layout change instead of this silently swallowing it.
NOT_RATED_TITLE_PREFIX = "Voti "
NOT_RATED_DISCLAIMER = ("Solo su www.fantacalcio.it i voti ufficiali per la tua lega",
                        "QUESTO FILE NON PUO' ESSERE RIPRODOTTO NE' PUBBLICATO SU ALTRI SITI INTERNET",
                        "E' DA CONSIDERARSI AD USO PERSONALE ESCLUSIVO DEGLI ISCRITTI DI FANTACALCIO.IT")


def _is_not_yet_rated(data: bytes) -> bool:
    """True when every sheet is exactly the title line plus the three fixed
    disclaimer lines, and nothing else -- see the ruling above."""
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            return False        # Finding F10: no sheets is a malformed file, never "not yet rated"
        for sheet in workbook.worksheets:
            rows = [[cell for cell in row if not _blank(cell)] for row in sheet.iter_rows(values_only=True)]
            if len(rows) != 1 + len(NOT_RATED_DISCLAIMER):
                return False
            title, *disclaimer = rows
            if len(title) != 1 or not str(title[0]).startswith(NOT_RATED_TITLE_PREFIX):
                return False
            if [d[0] if len(d) == 1 else None for d in disclaimer] != list(NOT_RATED_DISCLAIMER):
                return False
        return True
    finally:
        workbook.close()


def is_not_yet_rated_workbook(path: Path) -> bool:
    """The record-time twin of `_is_not_yet_rated`, for a file already on
    disk -- a workbook fetched before this ruling existed, or refetched
    after the site answered this shape again, is not re-derived from a live
    request, so this reads the same bytes back off `path`."""
    return _is_not_yet_rated(path.read_bytes())


@dataclass(frozen=True)
class VotoRow:
    sheet: str
    player_id: int
    name: str
    team: str
    classic_role: str
    voto: Decimal | None
    senza_voto: bool
    goals: int
    goals_conceded: int
    pen_saved: int
    pen_missed: int
    pen_scored: int
    own_goals: int
    yellow: int
    red: int
    assists: int
    raw: dict[str, Any]


@dataclass(frozen=True)
class VotiWorkbook:
    sheets: dict[str, list[VotoRow]]

    @property
    def rows(self) -> list[VotoRow]:
        return [row for rows in self.sheets.values() for row in rows]


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _jsonable(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def parse_voto(value: Any) -> tuple[Decimal | None, bool]:
    """(voto, senza_voto): a number, or None with the flag set for an unrated player."""
    if _blank(value):
        return None, True
    if isinstance(value, bool):
        raise VotiShapeError(f"unreadable voto {value!r}")
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
    else:
        text = str(value).strip().replace(",", ".")
        if text.endswith("*") or text.lower() in SENZA_VOTO_TEXT:
            return None, True
        try:
            number = Decimal(text)
        except InvalidOperation:
            raise VotiShapeError(f"unreadable voto {value!r}") from None
    if number == SENZA_VOTO_SENTINEL:
        return None, True
    return number, False


def _count(value: Any) -> int:
    if _blank(value):
        return 0
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except ValueError:
        raise VotiShapeError(f"unreadable event count {value!r}") from None


def _parse_sheet(sheet: Any, path: Path) -> list[VotoRow] | None:
    """The sheet's player rows, or None when it carries no voti table at all.

    The header row is not unique to the top of the sheet: the site repeats it
    before every club's block (observed in every one of the 20 blocks of
    captured/voti-21-01.xlsx and voti-20-01.xlsx, 2026-08-28) -- Task 6's
    probe only looked at the first block and missed the repeat. Every
    occurrence is validated and skipped, wherever it falls.

    A club-name row is tracked unconditionally, even before the sheet's own
    first header -- the club whose block opens the sheet has its name row
    *above* that header, and would otherwise never be recorded as `team`
    once header_seen flips true.
    """
    rows: list[VotoRow] = []
    header_seen = False
    team: str | None = None
    width = len(VOTI_HEADER)
    for index, values in enumerate(sheet.iter_rows(values_only=True)):
        cells = (list(values) + [None] * width)[:width]
        texts = [_text(c) for c in cells]
        if texts[:1] == [VOTI_HEADER[0]] and "Nome" in texts and "Voto" in texts:
            observed = tuple(t for t in texts if t)
            if observed != VOTI_HEADER:
                raise VotiShapeError(f"{path}: sheet {sheet.title!r}: header {observed} is not {VOTI_HEADER}")
            header_seen = True
            continue
        first, rest = cells[0], cells[1:]
        rest_blank = all(_blank(c) for c in rest)
        if isinstance(first, str) and first.strip() and rest_blank:
            team = first.strip()
            continue
        if not header_seen:
            if index >= HEADER_SEARCH_ROWS:
                return None
            continue
        if _blank(first) and rest_blank:
            continue
        try:
            player_id = int(first)
        except (TypeError, ValueError):
            raise VotiShapeError(f"{path}: sheet {sheet.title!r}: row {index + 1} is neither a club nor a player: {cells!r}") from None
        if team is None:
            raise VotiShapeError(f"{path}: sheet {sheet.title!r}: a player row before any club row")
        record = dict(zip(VOTI_HEADER, cells, strict=True))
        voto, senza_voto = parse_voto(record["Voto"])
        counts = {column: _count(record[header]) for header, column in COLUMNS.items()}
        rows.append(VotoRow(sheet=sheet.title, player_id=player_id, name=_text(record["Nome"]), team=team,
                            classic_role=_text(record["Ruolo"]), voto=voto, senza_voto=senza_voto,
                            raw={k: _jsonable(v) for k, v in record.items()}, **counts))
    return rows if header_seen else None


def parse_voti(path: Path) -> VotiWorkbook:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = {sheet.title: rows for sheet in workbook.worksheets
                  if (rows := _parse_sheet(sheet, path)) is not None}
    finally:
        workbook.close()
    if not sheets:
        raise VotiShapeError(f"{path}: no sheet carries the voti table (header {VOTI_HEADER})")
    return VotiWorkbook(sheets)


async def fetch_voti(http: httpx.AsyncClient, store: RawStore, *, cookie: str, season_id: int,
                     giornata: int) -> RawFile:
    url = VOTES_URL.format(season_id=season_id, giornata=giornata)
    data = await fetch_bytes(http, url, headers={"Cookie": cookie, "Accept": ACCEPT})
    if data[:4] != XLSX_MAGIC:
        if b"<html" in data[:2000].lower():
            # A login page with a 200 is the other way a dead session can look.
            raise WebSessionExpired(f"{url} -> HTTP 200 but an HTML page, not a workbook", url=url, status=200)
        raise VotiShapeError(f"{url}: not an xlsx ({len(data)} bytes)")
    if _is_not_published_placeholder(data):
        # Ruling R4: an unplayed giornata answers 200 with the placeholder
        # workbook, not a 404 -- treat it the same, and store nothing.
        raise NotPublished(f"{url} -> HTTP 200 but the placeholder workbook (not yet published)",
                           url=url, status=200)
    if _is_not_yet_rated(data):
        # Ruling R9: a giornata on the calendar that has not been rated yet
        # answers 200 with the title-and-disclaimer shell, not the
        # placeholder above -- treat it the same way: not published, store
        # nothing.
        raise NotPublished(f"{url} -> HTTP 200 but the not-yet-rated shell (no header, no rows)",
                           url=url, status=200)
    return store.write_bytes("voti", data, ext="xlsx", label=f"{season_id}-{giornata:02d}")


@dataclass(frozen=True)
class VotiFetch:
    raws: dict[int, RawFile]
    skipped: list[int]                  # already on disk and not --refetch
    not_published_from: int | None      # the first giornata the site has not published: a 404 or a placeholder workbook


async def fetch_voti_range(http: httpx.AsyncClient, store: RawStore, *, cookie: str, season_id: int,
                           giornate: Iterable[int], existing: set[int], refetch: bool = False) -> VotiFetch:
    """One season's workbooks, in order, one second apart, stopping at the
    first giornata the site has not published -- a 404, or a 200 carrying
    the placeholder workbook (Ruling R4): fetch_voti raises NotPublished for
    both, so this loop needs no extra branch.

    A giornata already on disk is skipped unless `refetch`: the files are
    immutable and the site republishes a giornata only to correct it, which
    is exactly what --refetch is for.
    """
    raws: dict[int, RawFile] = {}
    skipped: list[int] = []
    downloaded = 0
    for giornata in giornate:
        if giornata in existing and not refetch:
            skipped.append(giornata)
            continue
        if downloaded:
            await polite_pause()
        downloaded += 1
        try:
            raws[giornata] = await fetch_voti(http, store, cookie=cookie, season_id=season_id, giornata=giornata)
        except NotPublished:
            return VotiFetch(raws, skipped, giornata)
    return VotiFetch(raws, skipped, None)


@dataclass(frozen=True)
class VotiIngestResult:
    file_id: int | None
    season_id: int
    giornata: int
    inserted: int
    skipped_duplicate: bool
    sheets: list[str]
    unknown_players: int
    sha256: str
    raw_path: str

    def to_dict(self) -> dict[str, Any]:
        return {"file_id": self.file_id, "season_id": self.season_id, "giornata": self.giornata,
                "inserted": self.inserted, "skipped_duplicate": self.skipped_duplicate, "sheets": self.sheets,
                "unknown_players": self.unknown_players, "sha256": self.sha256, "raw_path": self.raw_path}


def record_voti(con: duckdb.DuckDBPyConnection, season_id: int, giornata: int, workbook: VotiWorkbook,
                raw: RawFile, *, known_ids: set[int]) -> VotiIngestResult:
    """Append one file row and its player rows; the same bytes twice for the
    same giornata is a no-op (the key is season, giornata *and* content: two
    giornate whose workbooks happen to be byte-identical are two files).

    `unknown_players` counts ids the current listone does not carry -- in a
    back season that is every player who has since left Serie A, so it is a
    count in the report rather than a row-by-row warning. Coach rows
    (`classic_role == COACH_ROLE`) are excluded from this count: a coach id
    is never in the listone, so counting it would always inflate "unknown"
    by one per club and bury the real signal (players signed after the
    listone snapshot). The coach row itself is still inserted below.
    """
    existing = con.execute(
        "SELECT file_id, sheets FROM voti_files WHERE season_id = ? AND giornata = ? AND sha256 = ?",
        [season_id, giornata, raw.sha256]).fetchone()
    if existing is not None:
        return VotiIngestResult(existing[0], season_id, giornata, 0, True, list(existing[1]), 0,
                                raw.sha256, str(raw.path))
    rows = workbook.rows
    unknown = {r.player_id for r in rows if r.classic_role != COACH_ROLE} - known_ids
    con.begin()
    try:
        file_id = con.execute(
            "INSERT INTO voti_files (season_id, giornata, fetched_at, source, raw_path, sha256, sheets, row_count) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?) RETURNING file_id",
            [season_id, giornata, to_db(raw.fetched_at), SOURCE, str(raw.path), raw.sha256,
             list(workbook.sheets), len(rows)]).fetchone()[0]
        con.executemany(
            "INSERT INTO player_match VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?::JSON)",
            [[file_id, season_id, giornata, r.sheet, r.player_id, r.name, r.team, r.classic_role, r.voto,
              r.senza_voto, r.goals, r.goals_conceded, r.pen_saved, r.pen_missed, r.pen_scored, r.own_goals,
              r.yellow, r.red, r.assists, json.dumps(r.raw, ensure_ascii=False)] for r in rows])
    except Exception:
        con.rollback()
        raise
    con.commit()
    return VotiIngestResult(file_id, season_id, giornata, len(rows), False, list(workbook.sheets),
                            len(unknown), raw.sha256, str(raw.path))
