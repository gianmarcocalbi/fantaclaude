import io
import json
from decimal import Decimal

import httpx
import openpyxl
import pytest
import respx
from fantaclaude.cli.app import ExitCode, app
from fantaclaude.commands.ingest import existing_giornate, record_voti_files
from fantaclaude.ingest.http import NotPublished, WebSessionExpired
from fantaclaude.ingest.listone_api import load_listone, record_listone
from fantaclaude.ingest.raw import RawStore
from fantaclaude.ingest.stats_web import (
    VOTES_URL,
    VOTI_HEADER,
    VotiFetch,
    VotiShapeError,
    fetch_voti,
    fetch_voti_range,
    is_not_yet_rated_workbook,
    parse_voti,
    parse_voto,
    record_voti,
)
from fantaclaude.league.settings import record_snapshot, snapshot_from_payloads
from typer.testing import CliRunner

COOKIE = "session=synthetic-value-for-tests; other=1"


@pytest.fixture
def no_pause(monkeypatch):
    async def fake(seconds=None):
        pass

    monkeypatch.setattr("fantaclaude.ingest.stats_web.polite_pause", fake)
    monkeypatch.setattr("fantaclaude.commands.ingest.polite_pause", fake)


@pytest.fixture
def sample_bytes(fixture_file):
    return fixture_file("voti_sample.xlsx").read_bytes()


@pytest.fixture
def placeholder_bytes(fixture_file):
    return fixture_file("voti_placeholder.xlsx").read_bytes()


@pytest.fixture
def not_yet_rated_bytes(fixture_file):
    return fixture_file("voti_not_yet_rated.xlsx").read_bytes()


def test_parse_voto_conventions():
    assert parse_voto(6.5) == (Decimal("6.5"), False)
    assert parse_voto(7) == (Decimal(7), False)
    assert parse_voto("6,5") == (Decimal("6.5"), False)
    assert parse_voto("6*") == (None, True)                 # voto d'ufficio: played, not rated
    assert parse_voto("s.v.") == (None, True) and parse_voto("S.V.") == (None, True)
    assert parse_voto(55) == (None, True) and parse_voto("55") == (None, True)   # fantacalcio.it's sentinel
    assert parse_voto(None) == (None, True) and parse_voto("  ") == (None, True)
    with pytest.raises(VotiShapeError):
        parse_voto("sette")


def test_parse_voti_reads_every_sheet_and_the_reference_players(fixture_file):
    workbook = parse_voti(fixture_file("voti_sample.xlsx"))
    assert workbook.sheets and all(rows for rows in workbook.sheets.values())
    scalvini_votes = set()
    for sheet, rows in workbook.sheets.items():
        by = {r.player_id: r for r in rows}
        assert {r.team.lower() for r in rows} == {"atalanta", "bologna"}
        assert all(r.sheet == sheet for r in rows)
        carnesecchi = by[4431]
        assert carnesecchi.name.lower().startswith("carnesecchi") and carnesecchi.classic_role.upper() == "P"
        assert (carnesecchi.voto, carnesecchi.senza_voto) == (Decimal("6.5"), False)
        assert (carnesecchi.goals_conceded, carnesecchi.assists, carnesecchi.goals) == (1, 1, 0)
        assert (by[4479].voto, by[4479].senza_voto) == (None, True)              # Elmas
        assert by[4371].voto == Decimal(7) and by[4371].goals == 1              # Raspadori
        assert by[6435].voto == Decimal(7) and by[6435].goals == 1              # Krstovic
        assert by[2640].voto == Decimal(6) and by[2640].classic_role.upper() == "D"   # Kolasinac
        assert set(by[4431].raw) == set(VOTI_HEADER)                              # the source row, as read
        scalvini_votes.add(by[5526].voto)
    assert scalvini_votes == {Decimal(6), Decimal("5.5")}
    assert len(workbook.rows) == sum(len(rows) for rows in workbook.sheets.values())


def test_parse_voti_fails_loud_on_an_appended_column(tmp_path, fixture_file):
    """A 14th column appended after `Ass` -- the likeliest way fantacalcio.it
    ever changes this shape -- must be caught the same as a renamed or
    reordered one. The old code truncated every row to len(VOTI_HEADER)
    cells *before* comparing the header text, so an appended column never
    reached the comparison at all and the parse silently succeeded with the
    extra data dropped. Built here from the committed fixture, the same way
    the reviewer proved the bug, never a hand-authored xlsx."""
    wb = openpyxl.load_workbook(fixture_file("voti_sample.xlsx"))
    sheet = wb.worksheets[0]
    extra_col = len(VOTI_HEADER) + 1
    for row in sheet.iter_rows():
        first = row[0].value
        is_header = first == VOTI_HEADER[0] and any(c.value == "Voto" for c in row)
        if is_header:
            sheet.cell(row=row[0].row, column=extra_col, value="Nuovo")
        elif isinstance(first, int):
            sheet.cell(row=row[0].row, column=extra_col, value=1)
    appended = tmp_path / "appended.xlsx"
    wb.save(appended)
    with pytest.raises(VotiShapeError, match="Nuovo"):
        parse_voti(appended)


def test_parse_voti_fails_loud_on_a_missing_club_row(tmp_path, fixture_file):
    """Deleting the club-name row that opens a block must not leave `team`
    silently carrying the previous block's club name forward -- the header
    row that follows must be immediately preceded by a club row, or this is
    a genuine layout surprise. Also covers the previously-deferred finding
    that a header immediately following another header (no club row, no
    player rows between them) keeps the same stale `team`."""
    wb = openpyxl.load_workbook(fixture_file("voti_sample.xlsx"))
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    bologna_row = next(i for i, r in enumerate(rows) if r and r[0] == "Bologna")
    assert rows[bologna_row + 1][0] == VOTI_HEADER[0]           # Bologna's own repeated header

    missing_club = tmp_path / "missing_club.xlsx"
    sheet.delete_rows(bologna_row + 1, 1)                        # 1-indexed: the "Bologna" row itself
    wb.save(missing_club)
    with pytest.raises(VotiShapeError, match="not preceded by a club row"):
        parse_voti(missing_club)


def test_parse_voti_fails_loud_on_a_header_immediately_after_a_header(tmp_path, fixture_file):
    wb = openpyxl.load_workbook(fixture_file("voti_sample.xlsx"))
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    atalanta_header = next(i for i, r in enumerate(rows) if r and r[0] == VOTI_HEADER[0])
    bologna_row = next(i for i, r in enumerate(rows) if r and r[0] == "Bologna")
    # Delete everything between Atalanta's header and Bologna's club row
    # (Atalanta's players), and the Bologna club row itself -- leaving
    # Atalanta's header immediately followed by Bologna's repeated header.
    start = atalanta_header + 2                                  # 1-indexed, first row after the header
    count = bologna_row - atalanta_header                        # players through the Bologna club row inclusive
    sheet.delete_rows(start, count)
    header_after_header = tmp_path / "header_after_header.xlsx"
    wb.save(header_after_header)
    with pytest.raises(VotiShapeError, match="not preceded by a club row"):
        parse_voti(header_after_header)


def test_parse_voti_fails_loud_on_layout_drift(tmp_path, fixture_file):
    wb = openpyxl.load_workbook(fixture_file("voti_sample.xlsx"))
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "Gf":
                cell.value = "Goal"
    renamed = tmp_path / "renamed.xlsx"
    wb.save(renamed)
    with pytest.raises(VotiShapeError, match="Goal"):
        parse_voti(renamed)
    empty = openpyxl.Workbook()
    empty.active.append(["not", "a", "voti", "table"])
    path = tmp_path / "empty.xlsx"
    empty.save(path)
    with pytest.raises(VotiShapeError, match="no sheet"):
        parse_voti(path)


def test_is_not_yet_rated_workbook_is_narrow(tmp_path, not_yet_rated_bytes):
    """Ruling R9: the not-yet-rated shell is matched exactly against the
    fixed title-and-disclaimer block -- one stray row beyond it, in any
    sheet, and it is not the shell any more. parse_voti must still fail
    loud on that near miss, the same as any other genuine layout change."""
    shell = tmp_path / "shell.xlsx"
    shell.write_bytes(not_yet_rated_bytes)
    assert is_not_yet_rated_workbook(shell)

    wb = openpyxl.load_workbook(shell)
    wb["Fantacalcio"].append(["something unexpected"])
    near_miss = tmp_path / "near_miss.xlsx"
    wb.save(near_miss)
    assert not is_not_yet_rated_workbook(near_miss)
    with pytest.raises(VotiShapeError, match="no sheet"):
        parse_voti(near_miss)


def test_is_not_yet_rated_workbook_rejects_a_workbook_with_no_sheets(tmp_path, monkeypatch):
    """Finding F10: a workbook with zero worksheets must not be classified as
    "not yet rated" -- the for loop over an empty worksheets list never
    runs, so the old code fell through to `return True` for any malformed
    file that happens to have no sheets at all. openpyxl refuses to *save* a
    workbook with no visible sheet (IndexError), so a truly empty one is
    only reachable via a corrupted or hand-crafted file -- simulated here by
    stubbing load_workbook rather than fighting that save-time guard."""
    class _NoSheets:
        worksheets: tuple = ()

        def close(self):
            pass

    monkeypatch.setattr("fantaclaude.ingest.stats_web.openpyxl.load_workbook", lambda *a, **k: _NoSheets())
    path = tmp_path / "no_sheets.xlsx"
    path.write_bytes(b"not a real workbook, load_workbook is stubbed above")
    assert not is_not_yet_rated_workbook(path)


def test_record_voti_files_counts_an_on_disk_shell_without_raising(db, tmp_path, not_yet_rated_bytes):
    """Ruling R9: a not-yet-rated workbook already on disk (fetched before
    this ruling existed, or refetched and still unrated) must be counted
    and reported by record_voti_files, never raised past it and never
    written to voti_files -- the same "fetched but unrecorded" divergence
    Ruling R8b closed, for a different cause. A near miss on the same disk
    must still raise, so the narrow match holds at the integration level too."""
    store = RawStore(tmp_path / "raw")
    store.write_bytes("voti", not_yet_rated_bytes, ext="xlsx", label="21-03")
    fetched = {21: VotiFetch(raws={}, skipped=[3], not_published_from=None)}
    results, not_yet_rated = record_voti_files(db, store, fetched, [3])
    assert results == [] and not_yet_rated == {21: [3]}
    assert db.execute("SELECT count(*) FROM voti_files").fetchone()[0] == 0

    wb = openpyxl.load_workbook(io.BytesIO(not_yet_rated_bytes))
    wb["Fantacalcio"].append(["something unexpected"])
    buf = io.BytesIO()
    wb.save(buf)
    store.write_bytes("voti", buf.getvalue(), ext="xlsx", label="21-04")
    with pytest.raises(VotiShapeError, match="no sheet"):
        record_voti_files(db, store, {21: VotiFetch(raws={}, skipped=[3, 4], not_published_from=None)}, [3, 4])


def test_existing_giornate_excludes_a_not_yet_rated_shell(tmp_path, sample_bytes, not_yet_rated_bytes):
    """Finding F3: existing_giornate must not count a not-yet-rated shell as
    "already on disk" -- fetch_voti_range treats existing_giornate's set as
    "do not re-download," so counting the shell would permanently suppress
    that giornata once the site actually rates it: nothing would ever ask
    for it again. record_voti_files (a separate call, on _voti_on_disk
    directly) still sees it and keeps reporting it as not-yet-rated."""
    store = RawStore(tmp_path / "raw")
    store.write_bytes("voti", sample_bytes, ext="xlsx", label="21-01")
    store.write_bytes("voti", sample_bytes, ext="xlsx", label="21-02")
    store.write_bytes("voti", not_yet_rated_bytes, ext="xlsx", label="21-03")
    assert existing_giornate(store, [21]) == {21: {1, 2}}                 # not 3


def _known(db, tmp_path, fixture_json) -> set[int]:
    raw = RawStore(tmp_path / "raw").write("listone", fixture_json("listone_sample"))
    record_listone(db, load_listone(raw.path), raw)
    return {r[0] for r in db.execute("SELECT player_id FROM v_players_current").fetchall()}


def test_record_voti_and_the_views(db, tmp_path, fixture_json, sample_bytes, fixture_file):
    known = _known(db, tmp_path, fixture_json)
    store = RawStore(tmp_path / "raw")
    raw = store.write_bytes("voti", sample_bytes, ext="xlsx", label="21-01")
    workbook = parse_voti(raw.path)
    result = record_voti(db, 21, 1, workbook, raw, known_ids=known)
    assert result.file_id == 1 and result.inserted == len(workbook.rows) and not result.skipped_duplicate
    assert result.sheets == list(workbook.sheets)
    # Ruling R6: the fixture's two coach rows (Sarri 688 for Atalanta, Tedesco
    # D. 5778 for Bologna, one pair per sheet -- 6 rows total) are stored like
    # any player row but excluded from unknown_players, which measures
    # listone coverage of players. Neither id is in the listone_sample.
    coach_ids = {r.player_id for r in workbook.rows if r.classic_role == "ALL"}
    non_coach_ids = {r.player_id for r in workbook.rows} - coach_ids
    assert coach_ids == {688, 5778}
    assert result.unknown_players == len(non_coach_ids - known) > 0
    assert result.unknown_players == len({r.player_id for r in workbook.rows} - known) - len(coach_ids)
    assert db.execute("SELECT count(*) FROM player_match WHERE classic_role = 'ALL'").fetchone()[0] == 6
    assert {r[0] for r in db.execute(
        "SELECT DISTINCT classic_role FROM player_match WHERE player_id IN (688, 5778)").fetchall()} == {"ALL"}
    assert db.execute("SELECT count(*) FROM v_player_match_current").fetchone()[0] == len(workbook.rows)
    season = db.execute("SELECT sheet, presenze, appearances, media_voto, goals FROM v_player_season "
                        "WHERE player_id = 2640 ORDER BY sheet").fetchall()
    assert len(season) == len(workbook.sheets) and all(row[1:] == (1, 1, Decimal("6.00"), 0) for row in season)
    form = db.execute("SELECT n, media_voto, last_giornata FROM v_player_form WHERE player_id = 2640").fetchall()
    assert form and form[0] == (1, Decimal("6.00"), 1)
    assert db.execute("SELECT senza_voto, voto FROM v_player_match_current WHERE player_id = 4479 LIMIT 1").fetchone() == (True, None)

    again = record_voti(db, 21, 1, workbook, raw, known_ids=known)
    assert again.skipped_duplicate and again.file_id == 1 and again.inserted == 0

    wb = openpyxl.load_workbook(fixture_file("voti_sample.xlsx"))
    wb.worksheets[0]["A1"] = "revised"
    revised = tmp_path / "revised.xlsx"
    wb.save(revised)
    raw2 = store.write_bytes("voti", revised.read_bytes(), ext="xlsx", label="21-01")
    second = record_voti(db, 21, 1, parse_voti(raw2.path), raw2, known_ids=known)
    assert second.file_id == 2
    assert db.execute("SELECT file_id FROM v_voti_files_current").fetchall() == [(2,)]
    assert db.execute("SELECT count(*) FROM player_match").fetchone()[0] == 2 * len(workbook.rows)   # history kept
    assert db.execute("SELECT count(*) FROM v_player_match_current").fetchone()[0] == len(workbook.rows)


@respx.mock
async def test_fetch_voti_sends_the_cookie_and_wants_a_workbook(tmp_path, sample_bytes, placeholder_bytes, not_yet_rated_bytes):
    url = VOTES_URL.format(season_id=21, giornata=1)
    route = respx.get(url).mock(return_value=httpx.Response(200, content=sample_bytes))
    async with httpx.AsyncClient() as http:
        raw = await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    assert raw.path.name.endswith("-voti-21-01.xlsx") and raw.path.read_bytes() == sample_bytes
    assert route.calls[0].request.headers["cookie"] == COOKIE
    respx.get(url).mock(return_value=httpx.Response(200, text="<html><body>Accedi</body></html>"))
    async with httpx.AsyncClient() as http:
        with pytest.raises(WebSessionExpired):
            await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    respx.get(url).mock(return_value=httpx.Response(200, content=b"garbage"))
    async with httpx.AsyncClient() as http:
        with pytest.raises(VotiShapeError):
            await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    respx.get(url).mock(return_value=httpx.Response(404))
    async with httpx.AsyncClient() as http:
        with pytest.raises(NotPublished):
            await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    # Ruling R4: an unplayed giornata answers 200 with the placeholder workbook,
    # not a 404 -- fetch_voti must treat it as NotPublished too, and write nothing.
    before = list((tmp_path / "raw" / "voti").glob("*.xlsx"))
    respx.get(url).mock(return_value=httpx.Response(200, content=placeholder_bytes))
    async with httpx.AsyncClient() as http:
        with pytest.raises(NotPublished):
            await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    assert list((tmp_path / "raw" / "voti").glob("*.xlsx")) == before
    # Ruling R9: a giornata that is on the calendar but has not been rated
    # yet answers 200 with the title-and-disclaimer shell, not the
    # placeholder above -- also NotPublished, also nothing written.
    respx.get(url).mock(return_value=httpx.Response(200, content=not_yet_rated_bytes))
    async with httpx.AsyncClient() as http:
        with pytest.raises(NotPublished):
            await fetch_voti(http, RawStore(tmp_path / "raw"), cookie=COOKIE, season_id=21, giornata=1)
    assert list((tmp_path / "raw" / "voti").glob("*.xlsx")) == before


@respx.mock
async def test_fetch_voti_range_skips_existing_and_stops_at_the_first_404(tmp_path, sample_bytes, monkeypatch):
    pauses = []

    async def count(seconds=None):
        pauses.append(seconds)

    monkeypatch.setattr("fantaclaude.ingest.stats_web.polite_pause", count)
    respx.get(url__regex=r".*/votes/21/(?P<g>\d+)$").mock(side_effect=lambda request, g: httpx.Response(
        200 if int(g) <= 2 else 404, content=sample_bytes if int(g) <= 2 else b""))
    store = RawStore(tmp_path / "raw")
    async with httpx.AsyncClient() as http:
        fetched = await fetch_voti_range(http, store, cookie=COOKIE, season_id=21, giornate=range(1, 39),
                                         existing={2}, refetch=False)
    assert sorted(fetched.raws) == [1] and fetched.skipped == [2] and fetched.not_published_from == 3
    assert len(pauses) == 1                                   # giornata 1, pause, giornata 3 (404)
    async with httpx.AsyncClient() as http:
        again = await fetch_voti_range(http, store, cookie=COOKIE, season_id=21, giornate=[1, 2],
                                       existing={1, 2}, refetch=True)
    assert sorted(again.raws) == [1, 2] and again.skipped == [] and again.not_published_from is None


@respx.mock
async def test_fetch_voti_range_stops_at_a_placeholder_workbook(tmp_path, sample_bytes, placeholder_bytes, no_pause):
    """Ruling R4: the site answers 200 with a placeholder workbook for a giornata
    that has not been played yet, not a 404 -- the range must stop there just the
    same, and must not write a raw file for the un-published giornata."""
    respx.get(url__regex=r".*/votes/21/(?P<g>\d+)$").mock(side_effect=lambda request, g: httpx.Response(
        200, content=sample_bytes if int(g) <= 2 else placeholder_bytes))
    store = RawStore(tmp_path / "raw")
    async with httpx.AsyncClient() as http:
        fetched = await fetch_voti_range(http, store, cookie=COOKIE, season_id=21, giornate=range(1, 39),
                                         existing=set(), refetch=False)
    assert sorted(fetched.raws) == [1, 2] and fetched.skipped == [] and fetched.not_published_from == 3
    assert list((tmp_path / "raw" / "voti").glob("*-voti-21-03.xlsx")) == []
    assert len(list((tmp_path / "raw" / "voti").glob("*.xlsx"))) == 2


def _seeded(tmp_path, fixture_json, mcp_fixture_json):
    from fantaclaude.db.connection import connect
    from fantaclaude.db.schema import apply_schema

    con = connect(tmp_path / "data" / "fanta.duckdb")
    apply_schema(con)
    raw = RawStore(tmp_path / "data" / "raw").write("listone", fixture_json("listone_sample"))
    record_listone(con, load_listone(raw.path), raw)
    record_snapshot(con, snapshot_from_payloads(
        profile=mcp_fixture_json("league_profile"), status=mcp_fixture_json("league_status"),
        rosters=mcp_fixture_json("roster_settings"), lineup=mcp_fixture_json("lineup_settings"),
        calculate=mcp_fixture_json("calculation_settings"), teams=mcp_fixture_json("teams")))
    con.close()


@respx.mock
def test_cli_ingest_stats_web(monkeypatch, tmp_path, fixture_json, mcp_fixture_json, sample_bytes, no_pause):
    monkeypatch.setenv("FANTACALCIO_HOME", str(tmp_path))
    monkeypatch.setenv("FANTACALCIO_WEB_COOKIE", COOKIE)
    _seeded(tmp_path, fixture_json, mcp_fixture_json)
    route = respx.get(url__regex=r".*/votes/(?P<s>\d+)/(?P<g>\d+)$").mock(side_effect=lambda request, s, g: httpx.Response(
        200 if int(g) <= 2 else 404, content=sample_bytes if int(g) <= 2 else b""))

    result = CliRunner().invoke(app, ["ingest", "stats-web", "--season", "21", "--json"])
    assert result.exit_code == ExitCode.OK, result.output
    payload = json.loads(result.stdout)["stats_web"]
    assert [(f["season_id"], f["giornata"], f["skipped_duplicate"]) for f in payload["files"]] == [(21, 1, False), (21, 2, False)]
    assert payload["skipped"] == {"21": []} and payload["not_published_from"] == {"21": 3}
    assert len(list((tmp_path / "data" / "raw" / "voti").glob("*-voti-21-*.xlsx"))) == 2
    assert existing_giornate(RawStore(tmp_path / "data" / "raw"), [21]) == {21: {1, 2}}

    again = CliRunner().invoke(app, ["ingest", "stats-web", "--season", "21"])
    assert again.exit_code == ExitCode.OK and "skipped 1-2" in again.stdout and "not published from 3" in again.stdout

    refetch = CliRunner().invoke(app, ["ingest", "stats-web", "--season", "21", "--giornata", "1", "--refetch", "--json"])
    assert refetch.exit_code == ExitCode.OK
    assert json.loads(refetch.stdout)["stats_web"]["files"][0]["skipped_duplicate"] is True

    bad = CliRunner().invoke(app, ["ingest", "stats-web", "--giornata", "40"])
    assert bad.exit_code == ExitCode.USAGE and "40" in bad.stderr

    route.mock(return_value=httpx.Response(401))                    # the same route: respx answers the first match
    expired = CliRunner().invoke(app, ["ingest", "stats-web", "--season", "20"])
    assert expired.exit_code == ExitCode.NOT_READY and "re-capture" in expired.stderr
    assert COOKIE not in expired.stderr and COOKIE not in expired.stdout

    monkeypatch.delenv("FANTACALCIO_WEB_COOKIE")
    missing = CliRunner().invoke(app, ["ingest", "stats-web"])
    assert missing.exit_code == ExitCode.NOT_READY and "FANTACALCIO_WEB_COOKIE" in missing.stderr
