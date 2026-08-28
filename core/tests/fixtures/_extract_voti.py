"""One-shot: build voti_sample.xlsx from captured/voti-21-01.xlsx (giornata 1, 2026-27),
voti_placeholder.xlsx from captured/voti-21-38.xlsx (the "not yet published" 404-style
placeholder), and voti_not_yet_rated.xlsx from captured/voti-21-03.xlsx (Ruling R9: a
third, distinct "not published" shape -- a giornata that is on the calendar but has not
been rated yet answers with the normal three sheet names and the normal title/disclaimer
block, then nothing else at all: no club row, no header, no data).

Run from the workspace root:  uv run python core/tests/fixtures/_extract_voti.py

Every sheet is kept, with everything above and including its header row and
then only the Atalanta and Bologna blocks (a team row followed by its player
rows), so the layout the parser locks -- title rows, header, team rows,
senza-voto cells -- is exactly the site's. This includes the header row the
site repeats before *every* club's block (all 20 of them, not just the
sheet's first) -- observed by dumping the full captured workbook, since
Task 6's probe only looked at the first block and missed the repeat; the
extracted fixture keeps both the sheet's opening header and Bologna's own
repeated one, so the parser is tested against the real shape. The reference
values the tests assert (Carnesecchi 6,5 with a goal conceded and an
assist, Elmas senza voto, Raspadori and Krstovic 7 with a goal, Scalvini 6 /
5,5 / 6 across the sources) were read off the public voti page on
2026-08-28. Values only: no styles, no formulas. Public voti, nothing to
scrub.

The placeholder is copied verbatim: one sheet, one cell, the site's own
"not yet published" text -- Ruling R4 of Task 7, since the site answers a
200 with this workbook for a giornata that has not been played yet, not a
404. Nothing to scrub there either.

voti_not_yet_rated.xlsx is also copied verbatim: three sheets, four rows
each (a title line naming the giornata, then the same three disclaimer
lines every workbook carries), nothing else. Public boilerplate, nothing to
scrub.
"""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
CAPTURE = ROOT / "captured" / "voti-21-01.xlsx"
OUT = Path(__file__).with_name("voti_sample.xlsx")
PLACEHOLDER_CAPTURE = ROOT / "captured" / "voti-21-38.xlsx"
PLACEHOLDER_OUT = Path(__file__).with_name("voti_placeholder.xlsx")
NOT_YET_RATED_CAPTURE = ROOT / "captured" / "voti-21-03.xlsx"
NOT_YET_RATED_OUT = Path(__file__).with_name("voti_not_yet_rated.xlsx")
TEAMS = {"atalanta", "bologna"}
HEADER_FIRST = "Cod."


def main() -> None:
    source = openpyxl.load_workbook(CAPTURE, read_only=True, data_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)
    kept = 0
    for sheet in source.worksheets:
        target = out.create_sheet(sheet.title)
        header_seen = False
        keep_block = False
        for row in sheet.iter_rows(values_only=True):
            cells = list(row)
            first, rest = cells[0] if cells else None, cells[1:] if cells else []
            blank_rest = all(c is None or (isinstance(c, str) and not c.strip()) for c in rest)
            # A team row's first cell is the club name and every other cell is
            # blank; a header row's first cell is "Cod." but its other cells
            # are not blank, so this never mistakes one for the other. Tracked
            # unconditionally -- even before the sheet's own first header --
            # so the club whose name row precedes that header (Atalanta) is
            # not missed once header_seen flips true.
            if isinstance(first, str) and first.strip() and blank_rest:
                keep_block = first.strip().lower() in TEAMS
            if not header_seen:
                target.append(cells)
                header_seen = isinstance(first, str) and first.strip() == HEADER_FIRST
                continue
            if keep_block and not (first is None and blank_rest):
                target.append(cells)
                kept += 1
        if not header_seen:
            out.remove(target)
    out.save(OUT)
    print(f"wrote {len(out.sheetnames)} sheet(s) {out.sheetnames}, {kept} rows, {OUT.stat().st_size} bytes")

    placeholder = openpyxl.load_workbook(PLACEHOLDER_CAPTURE, read_only=True, data_only=True)
    placeholder_out = openpyxl.Workbook()
    placeholder_out.remove(placeholder_out.active)
    for sheet in placeholder.worksheets:
        target = placeholder_out.create_sheet(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            target.append(list(row))
    placeholder_out.save(PLACEHOLDER_OUT)
    print(f"wrote {len(placeholder_out.sheetnames)} sheet(s) {placeholder_out.sheetnames}, "
         f"{PLACEHOLDER_OUT.stat().st_size} bytes")

    not_yet_rated = openpyxl.load_workbook(NOT_YET_RATED_CAPTURE, read_only=True, data_only=True)
    not_yet_rated_out = openpyxl.Workbook()
    not_yet_rated_out.remove(not_yet_rated_out.active)
    for sheet in not_yet_rated.worksheets:
        target = not_yet_rated_out.create_sheet(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            target.append(list(row))
    not_yet_rated_out.save(NOT_YET_RATED_OUT)
    print(f"wrote {len(not_yet_rated_out.sheetnames)} sheet(s) {not_yet_rated_out.sheetnames}, "
         f"{NOT_YET_RATED_OUT.stat().st_size} bytes")


if __name__ == "__main__":
    main()
